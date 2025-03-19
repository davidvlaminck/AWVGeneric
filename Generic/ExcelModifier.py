from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from API.EMInfraDomain import ApplicationEnum
from API.Enums import Environment


class ExcelModifier:
    def __init__(self, file_path: Path):
        self.file_path = file_path
        self.first_part_url_dict_eminfra = {
            Environment.PRD: 'https://apps.mow.vlaanderen.be/eminfra/assets/',
            Environment.TEI: 'https://apps-tei.mow.vlaanderen.be/eminfra/assets/',
            Environment.DEV: 'https://apps-dev.mow.vlaanderen.be/eminfra/assets/',
            Environment.AIM: 'https://apps-dev.mow.vlaanderen.be/eminfra/assets/'
        }
        self.first_part_url_dict_elisainfra = {
            Environment.PRD: 'https://services.apps.mow.vlaanderen.be/awvinfra/ui/?asset=',
            Environment.TEI: 'https://services.apps-tei.mow.vlaanderen.be/awvinfra/ui/?asset=',
            Environment.DEV: 'https://services.apps-dev.mow.vlaanderen.be/awvinfra/ui/?asset=',
            Environment.AIM: 'https://services-aim.apps-dev.mow.vlaanderen.be/awvinfra/ui/?asset='
        }

    def add_hyperlink(self, sheet: str = None, link_type: ApplicationEnum = ApplicationEnum.EM_INFRA,
                                env: Environment = Environment.PRD) -> None:
        """Adds a hyperlink to the Excel file

        Adds a hyperlink to the column named "uuid", linking to the applications eminfra or elisainfra.
        Cell is formatted (underline and text colour=blue)

        :param sheet: Excel sheet name
        :type sheet: str
        :param link_type: link to AWV application
        :type link_type: ApplicationEnum
        :param env: environment
        :type env: Environment
        :return: None
        """
        wb = load_workbook(filename=self.file_path)
        ws = wb[f"{sheet}"] if sheet else wb.active  # read active sheet when empty

        column_letter = next(
            (
                get_column_letter(col_idx)
                for col_idx, cell in enumerate(ws[1], start=1)
                if "uuid" in str.lower(cell.value)
            ),
            None,
        )

        for row in range(2, ws.max_row + 1):  # Data starts from row 2 in Excel
            asset_uuid = ws[f"{column_letter}{row}"].value
            if link_type == ApplicationEnum.ELISA_INFRA:
                url_prefix = self.first_part_url_dict_elisainfra[env]
            elif link_type == ApplicationEnum.EM_INFRA:
                url_prefix = self.first_part_url_dict_eminfra[env]
            else:
                raise ValueError("Value must be one of {'eminfra', 'elisainfra'}")
            # install a link
            ws[f"{column_letter}{row}"] = f'=HYPERLINK("{url_prefix}{asset_uuid}","{asset_uuid}")'
            # apply formatting
            ws[f"{column_letter}{row}"].font = Font(
                underline='single',
                color="0070C0")
        wb.save(self.file_path)