from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from API.Enums import Environment


class ExcelEditor:
    def __init__(self, file_path: Path):
        self.file_path = file_path
        self.first_part_url_dict_eminfra = {
            Environment.PRD: 'https://apps.mow.vlaanderen.be/eminfra/assets/',
            Environment.TEI: 'https://apps-tei.mow.vlaanderen.be/eminfra/assets/',
            Environment.DEV: 'https://apps-dev.mow.vlaanderen.be/eminfra/assets/',
            Environment.AIM: 'https://apps-dev.mow.vlaanderen.be/eminfra/assets/'
        }
        self.first_part_url_dict_elisainfra = {
            Environment.PRD: 'https://services.apps.mow.vlaanderen.be/awvinfra/ui/?asset=',
            Environment.TEI: 'https://services.apps-tei.mow.vlaanderen.be/awvinfra/ui/?asset=',
            Environment.DEV: 'https://services.apps-dev.mow.vlaanderen.be/awvinfra/ui/?asset=',
            Environment.AIM: 'https://services-aim.apps-dev.mow.vlaanderen.be/awvinfra/ui/?asset='
        }

    def convert_uuid_to_formula(self, sheet: str = None, link_type: str = 'eminfra',
                                env: Environment = Environment.PRD) -> None:
        """Convert uuid in Excel file to a formula
    
        Assumption that the column is called uuid (case-insensitive).

        :param sheet: str
            Name of the Excel sheet.
        :param link_type:
            eminfra or elisa
        :param env: Environment enumeration
        :return: None
        """
        wb = load_workbook(filename=self.file_path)
        ws = wb[f"{sheet}"] if sheet else wb.active  # read active sheet when empty

        # Find the column index for the column named "uuid"
        column_letter = None
        for col_idx, cell in enumerate(ws[1], start=1):  # Iterate through the first row (header)
            if str.lower(cell.value) == "uuid":  # Check if the column name matches "uuid"
                column_letter = get_column_letter(col_idx)
                break  # break out of the for-loop, once the column is found.

        # Write a formula in column "C" to sum columns "A" and "B"
        for row in range(2, ws.max_row + 1):  # Data starts from row 2 in Excel
            asset_uuid = ws[f"{column_letter}{row}"].value
            if link_type == 'eminfra':
                url_prefix = self.first_part_url_dict_eminfra[env]
            elif link_type == 'elisainfra':
                url_prefix = self.first_part_url_dict_elisainfra[env]
            else:
                raise ValueError("Value must be one of {'eminfra', 'elisainfra'}")
            # install a link
            ws[f"{column_letter}{row}"] = f'=HYPERLINK("{url_prefix}{asset_uuid}","{asset_uuid}")'
            # apply formatting
            ws[f"{column_letter}{row}"].font = Font(
                underline='single',
                color="0070C0")
        wb.save(self.file_path)